import pandas as pd
import openpyxl
from collections import defaultdict
from math import sqrt

import variable_formatation as vf
from parameters import DATA_VARS, ACURACIA
from file_operations import get_experiments, get_save_path, get_prompt_name, write_log
from value_comparation import comparisons

def fill_cell(df, row, col):
    cell = df.cell(row=row+2, column=col+1)
    cell.fill = openpyxl.styles.PatternFill(fgColor="FFFF00", fill_type = "solid")
    

# Receives 2 excel sheets and compares them
def compare(src: pd.DataFrame, dst: pd.DataFrame, name: str
            ) -> tuple[int, int, list[int], list[int]]:

    # receiving dataframes sizes
    num_rows1, num_cols1 = src.shape
    num_rows2, num_cols2 = dst.shape
    
    if num_cols1 != num_cols2:
        print('Número de colunas diferentes em', name,
              'src =', num_cols1, 'dst =', num_cols2)
        return -1, -1, [-1], [-1]

    if (num_rows1) != num_rows2:
        print('Número de linhas diferentes em', name,
                'src =',num_rows1, 'dst =', num_rows2)
        return -2, -2, [-2], [-2]

    total_errors = 0
    line_errors = 0
    col_errors = [0 for _ in range(num_cols1 - 1)]
    var_errors = dict()
    errors_per_line = [0 for i in range(num_rows1)]
    
    # opens the excel file to write the results
    writer = pd.ExcelWriter(name, engine='openpyxl')
    book = writer.book
    
    # Obtém a planilha para poder alterar cores
    sheet = book.active
    if sheet is None:
        sheet = book.create_sheet('Sheet1')

    # Percorre as células comparando valores
    names = src.columns
    for c in range(1, num_cols1):
        func = comparisons(names[c])
        var_errors[names[c]] = defaultdict(int)
        for r1 in range(num_rows1):
            v1 = src.iloc[r1, c]
            v2 = dst.iloc[r1, c]
            result, error = func(v1, v2)
            if type(result) == float:
                if type(var_errors[names[c]]) == defaultdict:
                    var_errors[names[c]] = 0
                var_errors[names[c]] += result
            else:
                var_errors[names[c]][result] += 1
            if error:
                fill_cell(sheet, r1, c)
                col_errors[c-1] += 1
                errors_per_line[r1] += 1
                total_errors += 1
                if errors_per_line[r1] == 1:
                    fill_cell(sheet, r1, 0)
                    line_errors += 1
        if type(var_errors[names[c]]) == defaultdict:
            if 'TP' not in var_errors[names[c]]:
                var_errors[names[c]]['TP'] = 0
            if 'TN' not in var_errors[names[c]]:
                var_errors[names[c]]['TN'] = 0
            if 'FP' not in var_errors[names[c]]:
                var_errors[names[c]]['FP'] = 0
            if 'FN' not in var_errors[names[c]]:
                var_errors[names[c]]['FN'] = 0
            

    # Exporta o dataframe modificado
    dst.to_excel(writer, index=False) 
    writer._save()

    return total_errors, line_errors, col_errors, errors_per_line, var_errors

def float_string(num):
    num_int = int(num)
    dec_num = round(num - num_int, 2)*100
    if dec_num == 100:
        dec_num = 0
        num += 1
    return ('%.2d,%.2d' % (num, dec_num))

def get_percentage(value, total):
    num = (1 - (value/total)) * 100
    return float_string(num) + '%'

def measure_results(total_errors, sentence_errors, errors_per_col,
                    errors_per_line, var_errors, n_sentences,
                    n_variaveis, variaveis):
    total_values = n_sentences * n_variaveis
    ac_total = get_percentage(total_errors, total_values)
    ac_total_sent = get_percentage(sentence_errors, n_sentences)
    csv_line_f1 = [ac_total, ac_total_sent]
    csv_line_cm = [ac_total, ac_total_sent]

    output = f"Nº de Sentenças: {n_sentences} | Nº de Variáveis: {n_variaveis} | Nº Total de Valores: {total_values}\n"

    # Imprime resultados
    output += ('\nAcurácia por Variável:\n')
    var_idx = 0
    for i in range(n_variaveis):
        var_idx += 1
        while variaveis[i] != DATA_VARS[var_idx]:
            # se o prompt não contiver uma variável dentre as esperadas,
            # adiciona um valor nulo, para acurácia e F1-Score
            substituto = ['////', '////']
            csv_line_f1.extend(substituto)
            csv_line_cm.extend(substituto)
            if 'intervalo' not in DATA_VARS[var_idx]:
                # se não for uma variável de intervalo,
                # adiciona os valores substitutos de TP, TN, FP, FN
                csv_line_cm.extend(substituto + substituto)
            var_idx += 1
        ac = get_percentage(errors_per_col[i], n_sentences)
        csv_line_f1.append(ac)
        csv_line_cm.append(ac)
        resultado = 'Acertos: ' + ac
        log = var_errors[variaveis[i]]
        log_str = ' |'
        if type(log) != float:
            den = log['TP'] + log['FP']
            precision = (log['TP']/den) if den != 0 else 0
            den = log['TP'] + log['FN']
            recall = (log['TP']/den) if den != 0 else 0
            den = precision + recall
            f1 = (2*(precision*recall)/den) if den != 0 else 0
            f1 = float_string(f1)
            csv_line_f1.append(f1)
            csv_line_cm.append(f1)
            keys = list(log.keys())
            keys.sort()
            keys.reverse()
            for key in keys:
                val = get_percentage(abs(n_sentences - log[key]), n_sentences)
                str_log = str(val).rjust(3)
                csv_line_cm.append(val)
                str_key = str(key).rjust(2)
                log_str += ' ' + str_key + ': ' + str_log + ' |'
        else:
            log = sqrt(log/n_sentences)
            log = float_string(log)
            csv_line_f1.append(log)
            csv_line_cm.append(log)
            log_str += ' RMSN (horas): ' + log + ' |'
        # formats the variable name to fit 42 caracteres, filling with white spaces
        output += ' ' + variaveis[i].ljust(41) + ' : ' + resultado + log_str + '\n'

    output += ("-------------------------------------------\n")
    output += ('>> Acurácia Total das Sentenças: %s\n' % ac_total_sent)
    output += (">> Acurácia Total: %s\n" % ac_total)
    return csv_line_f1, csv_line_cm, output

def main(source, teste, saida_excel):
    # Lê arquivos
    df1 = vf.format_data(source)
    df2 = vf.format_data(teste)
    
    # Compara arquivos
    comparison = compare(df1, df2, saida_excel)

    if comparison[0] < 0:
        return -1, -1, -1
    n_sentences= df1.shape[0]
    # Por causa da coluna da sentença
    variaveis = df1.columns[1:]
    n_variaveis = len(variaveis)
    return measure_results(*comparison, n_sentences, n_variaveis,
                  variaveis)

def run_comparisons():
    experimentos_f1 = []
    experimentos_cm = []
    exp = get_experiments()
    for src, teste in exp:
        saida_excel, log_file = get_save_path(teste)
        csv_line_f1, csv_line_cm, log = main(src, teste, saida_excel)
        if log == -1:
            continue
        write_log(log_file, log)
        name = get_prompt_name(teste)
        experimentos_f1.append([name, '-------'] + csv_line_f1)
        experimentos_cm.append([name, '-------'] + csv_line_cm)
    
    vars_f1 = []
    vars_cm = []
    for i in DATA_VARS[1:]:
        vars_f1.append('AC - ' + i)
        vars_cm.append('AC - ' + i)
        if 'intervalo' in i:
            vars_f1.append('RMSN (h) - ' + i)
            vars_cm.append('RMSN (h) - ' + i)
        else:
            vars_f1.append('F1 - ' + i)
            vars_cm.append('F1 - ' + i)
            vars_cm.append('TP - ' + i)
            vars_cm.append('TN - ' + i)
            vars_cm.append('FP - ' + i)
            vars_cm.append('FN - ' + i)
    
    csv_header_f1 = ['Prompt', 'Descrição', 'Acurácia Total', 'Acurácia por Sentença'] + vars_f1
    csv_header_cm = ['Prompt', 'Descrição', 'Acurácia Total', 'Acurácia por Sentença'] + vars_cm
    pd.DataFrame(experimentos_f1, columns=csv_header_f1).to_csv(ACURACIA, index=False)
    pd.DataFrame(experimentos_cm, columns=csv_header_cm).to_csv(ACURACIA.replace('F1-Score', 'ConfusionMatrix'), index=False)

if __name__ == '__main__':
    run_comparisons()

